from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

class PresupuestoExcelReport:

    def __init__(
        self,
        presupuesto,
        capitulos,
        items
    ):
        self.presupuesto = presupuesto
        self.capitulos = capitulos
        self.items = items

    def generate(self, filename):

        workbook = Workbook()

        self.create_summary_sheet(workbook)
        self.create_chapters_sheet(workbook)
        self.create_items_sheet(workbook)

        workbook.save(filename)

        return filename

    def title_style(self, cell):

        cell.font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    def money(self, value):
        try:
            return (
                f"${float(value or 0):,.2f}"
                .replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            )
        except:
            return "$0,00"

    def create_summary_sheet(self, workbook):

        ws = workbook.active
        ws.title = "Resumen"
        ws["A1"] = "PRESUPUESTO BUILDORA"
        self.title_style(ws["A1"])         

        ws["A3"] = "Código"
        ws["B3"] = self.presupuesto["codigo_presupuesto"]

        ws["A4"] = "Nombre"
        ws["B4"] = self.presupuesto["nombre_presupuesto"]

        ws["A5"] = "Proyecto"
        ws["B5"] = self.presupuesto["nombre_proyecto"]

        ws["A6"] = "Cliente"
        ws["B6"] = self.presupuesto["nombre_cliente"]

        ws["A7"] = "Fecha"
        ws["B7"] = str(
            self.presupuesto["fecha_presupuesto"]
        )      
        ws["B10"] = self.presupuesto["costo_directo_total"]
        ws["A10"] = "Costo Directo"
        ws["B10"] = self.money(
            self.presupuesto["costo_directo_total"]
        )

        ws["A11"] = "Administración"
        ws["B11"] = self.money(
            self.presupuesto["total_administracion"]
        )

        ws["A12"] = "Imprevistos"
        ws["B12"] = self.money(
            self.presupuesto["total_imprevistos"]
        )

        ws["A13"] = "Utilidad"
        ws["B13"] = self.money(
            self.presupuesto["total_utilidad"]
        )

        ws["A14"] = "IVA Utilidad"
        ws["B14"] = self.money(
            self.presupuesto["iva_utilidad"]
        )

        ws["A16"] = "VALOR TOTAL"
        ws["B16"] = self.money(
            self.presupuesto["valor_total_presupuesto"]
        )

    def create_chapters_sheet(self, workbook):

        ws = workbook.create_sheet("Capitulos")

        headers = [
            "Código",
            "Nombre",
            "Costo",
            "Participación %"
        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col
            )

            cell.value = header

            self.title_style(cell)

        row = 2

        for chapter in self.capitulos:

            ws.cell(row,1).value = chapter["codigo_capitulo"]
            ws.cell(row,2).value = chapter["nombre_capitulo"]
            ws.cell(row,3).value = self.money(
                chapter["costo_total_capitulo"]
            )
            ws.cell(row,4).value = chapter["porcentaje_participacion"]

            row += 1

    def create_items_sheet(self, workbook):

        ws = workbook.create_sheet("Items")
            
        headers = [
            "Capítulo",
            "Código Ítem",
            "Descripción",
            "Unidad",
            "Cantidad",
            "Vr Unitario",
            "Vr Total",
            "APU"
        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col
            )

            cell.value = header

            self.title_style(cell)

        row = 2

        for item in self.items:

            ws.cell(row,1).value = item["nombre_capitulo"]
            ws.cell(row,2).value = item["codigo_item"]
            ws.cell(row,3).value = item["descripcion_item"]
            ws.cell(row,4).value = item["unidad_item"]
            ws.cell(row,5).value = item["cantidad"]
            ws.cell(row,6).value = self.money(
                item["costo_unitario_historico"]
            )
            ws.cell(row,7).value = self.money(
                item["costo_total_item"]
            )
            ws.cell(row,8).value = item["codigo_apu_historico"]

            row += 1